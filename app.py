import glob
import os
from openpyxl import load_workbook
from os import path
from database import Database

# excel files directory
dir_path = ''

# read/write databese log info or create it
if not os.path.isfile('log_info.txt'):
    log_file = open('log_info.txt', 'w')
    log_file.write('user: '     + input('enter user: ')     + '\n')
    log_file.write('password: ' + input('enter password: ') + '\n')
    log_file.write('database: ' + input('enter database: ') + '\n')
    log_file.write('host: '     + input('enter host: '))
    log_file.close()
log_file       = open('log_info.txt', 'r')
log_file_lines = log_file.readlines()
log_file.close()

user     = log_file_lines[0].split(' ')[len(log_file_lines[0].split(' ')) - 1]
user     = user[0:len(user) - 1]
password = log_file_lines[1].split(' ')[len(log_file_lines[1].split(' ')) - 1]
password = password[0:len(password) - 1]
database = log_file_lines[2].split(' ')[len(log_file_lines[2].split(' ')) - 1]
database = database[0:len(database) - 1]
host     = log_file_lines[3].split(' ')[len(log_file_lines[3].split(' ')) - 1]
host     = host[0:len(host)]

#connect to database
database = Database(user, password, database, host)
database.connect()

#iterate over .xlsx files
files = glob.glob(dir_path + '*.xlsx')
for file in files:

    if len(file) >= 2 and file[0:2] == '~$': #continue if .xlsx file is opened file
        continue

    cur_workbook = load_workbook(filename=file)
    cur_sheet = cur_workbook.active
    table_name = ''
    #iterate over rows
    for cur_row in cur_sheet.iter_rows():
        if cur_row[0].row == 1:
            if cur_row[2].value == 'ფართობი':
                table_name = 'area'
                if not database.table_exists(table_name):
                    database.create_table('ethernet_code', 'country_name', table_name, 'file_name', 'integer')
            else:
                table_name = 'capital'
                if not database.table_exists(table_name):
                    database.create_table('ethernet_code', 'country_name', table_name, 'file_name', 'character varying(100)')
        if cur_row[0].fill.start_color.rgb != '00000000' and cur_row[0].row != 1:
            file_name     = '\'' + file + '\''
            ethernet_code = '\'' + cur_row[0].value + '\''
            country_name  = '\'' + cur_row[1].value + '\''
            third_param   = str(cur_row[2].value)
            if not third_param.isdigit():
                third_param = '\'' + third_param + '\''
            database.insert_row(table_name,
                                'ethernet_code', 'country_name', table_name, 'file_name',
                                ethernet_code, country_name, third_param, file_name)

database.disconnect()
print('Success! Excel data is in database.')
